# ============================================================
# Motor de Cálculo BT — Versión Final
# Mejora A: Sugerencia automática de conductor mínimo
# Mejora B: Exportar resultados a Excel formateado
# Normativa: SEC RIC N°10 / NEC / IEC 60364
# ============================================================

from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- CONSTANTES ---
RHO_CU    = 0.0175
LIMITE_DV = 3.0

# --- SISTEMAS ---
TENSION_SISTEMA = {"3F": 380, "1F": 220, "2F": 220}
FACTOR_SISTEMA  = {"3F": 1.732, "1F": 2.0, "2F": 2.0}

# --- CONDUCTORES ---
# Orden importa — de menor a mayor sección
# La sugerencia recorre la lista y encuentra el primero que cumple
CONDUCTORES = {
    "14AWG":  {"mm2": 2.08,  "I_max": 20},
    "12AWG":  {"mm2": 3.31,  "I_max": 25},
    "10AWG":  {"mm2": 5.26,  "I_max": 35},
    "8AWG":   {"mm2": 8.37,  "I_max": 50},
    "6AWG":   {"mm2": 13.3,  "I_max": 65},
    "4AWG":   {"mm2": 21.1,  "I_max": 85},
    "2AWG":   {"mm2": 33.6,  "I_max": 115},
    "1/0AWG": {"mm2": 53.5,  "I_max": 150},
    "2/0AWG": {"mm2": 67.4,  "I_max": 175},
    "4/0AWG": {"mm2": 107.0, "I_max": 230},
    "350MCM": {"mm2": 177.0, "I_max": 310},
    "400MCM": {"mm2": 203.0, "I_max": 335},
    "500MCM": {"mm2": 253.0, "I_max": 380},
}

# --- TEMPERATURA ---
FACTORES_TEMP = {
    25: 1.04, 30: 1.00, 35: 0.96,
    40: 0.91, 45: 0.87, 50: 0.82,
}

# ============================================================
# FUNCIONES DE CÁLCULO
# ============================================================

def factor_temperatura(temp_amb):
    return FACTORES_TEMP.get(int(temp_amb), 1.00)

def capacidad_corregida(I_max, paralelos, temp_amb):
    return round(I_max * paralelos * factor_temperatura(temp_amb), 1)

def calcular_potencia(I_diseno, cos_phi, sistema):
    V = TENSION_SISTEMA.get(sistema, 380)
    if sistema == "3F":
        return round(1.732 * V * I_diseno * cos_phi)
    return round(V * I_diseno * cos_phi)

def calcular_caida_tension(L_m, S_mm2, I_diseno, paralelos, sistema):
    S_eq   = S_mm2 * paralelos
    factor = FACTOR_SISTEMA.get(sistema, 1.732)
    V_nom  = TENSION_SISTEMA.get(sistema, 380)
    dV_V   = (factor * RHO_CU * L_m * I_diseno) / S_eq
    dV_pct = (dV_V / V_nom) * 100
    return round(dV_V, 3), round(dV_pct, 3)

def clasificar_caida(dV_pct):
    if dV_pct <= 1.5:
        return "ÓPTIMO"
    elif dV_pct <= 3.0:
        return "ACEPTABLE"
    elif dV_pct <= 5.0:
        return "PRECAUCIÓN"
    else:
        return "FALLA"

# ============================================================
# MEJORA A — SUGERENCIA AUTOMÁTICA DE CONDUCTOR MÍNIMO
# ============================================================

def sugerir_conductor(L_m, I_diseno, paralelos, sistema, temp_amb):
    """
    Recorre la tabla de conductores de menor a mayor sección.
    Retorna el primero que cumple AMBAS condiciones:
    1. Caída de tensión <= límite normativo
    2. Capacidad de corriente >= corriente de diseño
    Si ninguno cumple retorna None.
    """
    for nombre, datos in CONDUCTORES.items():
        S_mm2 = datos["mm2"]
        I_max = datos["I_max"]

        # Verificar caída de tensión
        _, dV_pct = calcular_caida_tension(L_m, S_mm2, I_diseno, paralelos, sistema)

        # Verificar capacidad con corrección de temperatura
        I_cap = capacidad_corregida(I_max, paralelos, temp_amb)

        # Ambas condiciones deben cumplirse
        if dV_pct <= LIMITE_DV and I_cap >= I_diseno:
            return nombre, S_mm2, round(dV_pct, 3)

    # Ningún conductor en la tabla es suficiente
    return None, None, None

# ============================================================
# LECTURA EXCEL
# ============================================================

def leer_circuitos_excel(nombre_archivo):
    circuitos = []
    errores   = []

    libro = openpyxl.load_workbook(nombre_archivo, data_only=True)
    hoja  = libro.active

    print(f"\n  Leyendo: {nombre_archivo}")
    print(f"  Circuitos encontrados: {hoja.max_row - 1}")

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        nombre    = fila[0]
        sistema   = fila[1]
        conductor = fila[2]
        paralelos = fila[3]
        I_diseno  = fila[4]
        cos_phi   = fila[5]
        L_m       = fila[6]
        temp_amb  = fila[7]

        if nombre is None:
            continue

        sistema   = str(sistema).strip().upper()
        conductor = str(conductor).strip().upper()

        if sistema not in FACTOR_SISTEMA:
            errores.append(f"'{nombre}': sistema '{sistema}' inválido")
            continue

        if conductor not in CONDUCTORES:
            errores.append(f"'{nombre}': conductor '{conductor}' no existe")
            continue

        circuitos.append({
            "nombre":    str(nombre).strip(),
            "sistema":   sistema,
            "conductor": conductor,
            "S_mm2":     CONDUCTORES[conductor]["mm2"],
            "I_max":     CONDUCTORES[conductor]["I_max"],
            "paralelos": int(paralelos),
            "I_diseno":  float(I_diseno),
            "cos_phi":   float(cos_phi),
            "L_m":       float(L_m),
            "temp_amb":  float(temp_amb),
        })

    if errores:
        print("\n  ADVERTENCIAS:")
        for e in errores:
            print(f"  ⚠ {e}")

    return circuitos

# ============================================================
# MEJORA B — EXPORTAR A EXCEL FORMATEADO
# ============================================================

def exportar_excel(nombre_proyecto, circuitos, fecha, nombre_archivo):
    """
    Exporta resultados a Excel con formato profesional.
    Colores por estado: verde ÓPTIMO, amarillo ACEPTABLE,
    naranja PRECAUCIÓN, rojo FALLA.
    """

    # Colores por estado — formato ARGB (Alpha+RGB)
    COLORES = {
        "ÓPTIMO":     "FF92D050",   # verde
        "ACEPTABLE":  "FFFFFF00",   # amarillo
        "PRECAUCIÓN": "FFFFC000",   # naranja
        "FALLA":      "FFFF0000",   # rojo
        "SUPERA":     "FFFF0000",   # rojo para corriente
    }

    # Estilo de borde para las celdas
    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    libro = openpyxl.Workbook()
    hoja  = libro.active
    hoja.title = "Resultados"

    # --- TÍTULO DEL PROYECTO ---
    hoja.merge_cells("A1:K1")
    hoja["A1"] = f"REPORTE BT — {nombre_proyecto}"
    hoja["A1"].font      = Font(bold=True, size=14)
    hoja["A1"].alignment = Alignment(horizontal="center")
    hoja["A1"].fill      = PatternFill("solid", fgColor="FF1F3864")
    hoja["A1"].font      = Font(bold=True, size=13, color="FFFFFFFF")

    hoja.merge_cells("A2:K2")
    hoja["A2"] = f"Fecha: {fecha}  |  Normativa: SEC RIC N°10 / NEC / IEC 60364  |  Límite ΔV: {LIMITE_DV}%"
    hoja["A2"].alignment = Alignment(horizontal="center")
    hoja["A2"].fill      = PatternFill("solid", fgColor="FFD6E4F7")

    # --- ENCABEZADOS ---
    encabezados = [
        "Circuito", "Sistema", "Conductor", "Paralelos",
        "I_diseño(A)", "I_cap(A)", "Estado_I",
        "Potencia(W)", "ΔV(V)", "ΔV(%)", "Estado_dV"
    ]

    fila_enc = 3
    for col, texto in enumerate(encabezados, start=1):
        celda = hoja.cell(row=fila_enc, column=col, value=texto)
        celda.font      = Font(bold=True, color="FFFFFFFF")
        celda.fill      = PatternFill("solid", fgColor="FF2E75B6")
        celda.alignment = Alignment(horizontal="center")
        celda.border    = borde

    # --- DATOS ---
    for i, c in enumerate(circuitos):
        fila_num = fila_enc + 1 + i

        # Calcular
        I_cap        = capacidad_corregida(c["I_max"], c["paralelos"], c["temp_amb"])
        P_watts      = calcular_potencia(c["I_diseno"], c["cos_phi"], c["sistema"])
        dV_V, dV_pct = calcular_caida_tension(
                           c["L_m"], c["S_mm2"], c["I_diseno"],
                           c["paralelos"], c["sistema"]
                       )
        estado_dV    = clasificar_caida(dV_pct)
        estado_I     = "OK" if c["I_diseno"] <= I_cap else "SUPERA"

        # Sugerencia si hay falla
        sugerencia = ""
        if estado_dV == "FALLA" or estado_I == "SUPERA":
            cond_sug, mm2_sug, dv_sug = sugerir_conductor(
                c["L_m"], c["I_diseno"], c["paralelos"],
                c["sistema"], c["temp_amb"]
            )
            if cond_sug:
                sugerencia = f"→ Usar {cond_sug} ({dv_sug}%)"

        # Conductor con paralelos
        if c["paralelos"] > 1:
            desc_cond = f"{c['paralelos']}x{c['conductor']}"
        else:
            desc_cond = c["conductor"]

        # Escribir fila
        valores = [
            c["nombre"], c["sistema"], desc_cond, c["paralelos"],
            c["I_diseno"], I_cap, estado_I,
            P_watts, dV_V, dV_pct, estado_dV
        ]

        for col, valor in enumerate(valores, start=1):
            celda = hoja.cell(row=fila_num, column=col, value=valor)
            celda.border    = borde
            celda.alignment = Alignment(horizontal="center")

        # Color según estado de corriente
        color_I = COLORES.get(estado_I, "FFFFFFFF")
        hoja.cell(row=fila_num, column=7).fill = PatternFill("solid", fgColor=color_I)

        # Color según estado de caída
        color_dV = COLORES.get(estado_dV, "FFFFFFFF")
        hoja.cell(row=fila_num, column=11).fill = PatternFill("solid", fgColor=color_dV)

        # Sugerencia en columna L si hay falla
        if sugerencia:
            celda_sug = hoja.cell(row=fila_num, column=12, value=sugerencia)
            celda_sug.font   = Font(bold=True, color="FFFF0000")
            celda_sug.border = borde

    # Encabezado columna sugerencia
    hoja.cell(row=fila_enc, column=12, value="Sugerencia").font = Font(bold=True, color="FFFFFFFF")
    hoja.cell(row=fila_enc, column=12).fill   = PatternFill("solid", fgColor="FF2E75B6")
    hoja.cell(row=fila_enc, column=12).border = borde

    # --- AJUSTAR ANCHO DE COLUMNAS ---
    anchos = [28, 8, 14, 10, 12, 10, 10, 12, 8, 8, 12, 20]
    for col, ancho in enumerate(anchos, start=1):
        hoja.column_dimensions[
            openpyxl.utils.get_column_letter(col)
        ].width = ancho

    libro.save(nombre_archivo)
    print(f"  Excel guardado  : {nombre_archivo}")

# ============================================================
# REPORTE TXT
# ============================================================

def generar_reporte_txt(nombre_proyecto, circuitos, fecha):
    lineas      = []
    total_ok    = 0
    total_falla = 0

    lineas.append("=" * 60)
    lineas.append(f"  REPORTE — {nombre_proyecto}")
    lineas.append(f"  Fecha        : {fecha}")
    lineas.append(f"  Normativa    : SEC RIC N10 / NEC / IEC 60364")
    lineas.append(f"  Limite caida : {LIMITE_DV}% circuito final / 5% total")
    lineas.append(f"  Circuitos    : {len(circuitos)}")
    lineas.append("=" * 60)

    for c in circuitos:
        I_cap        = capacidad_corregida(c["I_max"], c["paralelos"], c["temp_amb"])
        P_watts      = calcular_potencia(c["I_diseno"], c["cos_phi"], c["sistema"])
        dV_V, dV_pct = calcular_caida_tension(
                           c["L_m"], c["S_mm2"], c["I_diseno"],
                           c["paralelos"], c["sistema"]
                       )
        estado_dV = clasificar_caida(dV_pct)
        estado_I  = "OK" if c["I_diseno"] <= I_cap else "SUPERA"

        if c["paralelos"] > 1:
            desc_cond = f"{c['paralelos']}x{c['conductor']} (S={c['S_mm2']*c['paralelos']}mm2)"
        else:
            desc_cond = f"{c['conductor']} ({c['S_mm2']}mm2)"

        V_nom = TENSION_SISTEMA[c["sistema"]]
        lineas.append("")
        lineas.append(f"  Circuito  : {c['nombre']}")
        lineas.append(f"  Sistema   : {c['sistema']} / {V_nom}V | Temp: {c['temp_amb']}C")
        lineas.append(f"  Conductor : {desc_cond}")
        lineas.append(f"  Corriente : {c['I_diseno']}A -> {estado_I} (cap. {I_cap}A)")
        lineas.append(f"  Potencia  : {P_watts} W")
        lineas.append(f"  Caida dV  : {dV_V}V ({dV_pct}%) -> {estado_dV}")

        # Sugerencia automática si hay falla
        if estado_dV == "FALLA" or estado_I == "SUPERA":
            cond_sug, mm2_sug, dv_sug = sugerir_conductor(
                c["L_m"], c["I_diseno"], c["paralelos"],
                c["sistema"], c["temp_amb"]
            )
            if cond_sug:
                lineas.append(f"  SUGERENCIA: usar {cond_sug} ({mm2_sug}mm2) -> dV={dv_sug}%")
            else:
                lineas.append(f"  SUGERENCIA: ningún conductor en tabla es suficiente")
            total_falla += 1
        else:
            total_ok += 1

    lineas.append("")
    lineas.append("=" * 60)
    lineas.append(f"  Circuitos OK    : {total_ok}")
    lineas.append(f"  Circuitos FALLA : {total_falla}")
    lineas.append("=" * 60)

    return lineas, total_ok, total_falla

def guardar_txt(lineas, nombre_archivo):
    with open(nombre_archivo, "w", encoding="utf-8") as f:
        for linea in lineas:
            f.write(linea + "\n")
    print(f"  TXT guardado    : {nombre_archivo}")

# ============================================================
# PROGRAMA PRINCIPAL
# ============================================================

print("=" * 60)
print("  MOTOR DE CALCULO BT — VERSION FINAL")
print("  1F / 2F / 3F | Paralelos | Temp | Sugerencia")
print("  Normativa: SEC RIC N10 / NEC / IEC 60364")
print("=" * 60)

ahora         = datetime.now()
fecha         = ahora.strftime("%d/%m/%Y %H:%M")
fecha_archivo = ahora.strftime("%Y%m%d_%H%M")

nombre_proyecto = input("\n  Nombre del proyecto : ").strip()
archivo_excel   = input("  Archivo Excel       : ").strip()

if not archivo_excel.endswith(".xlsx"):
    archivo_excel += ".xlsx"

try:
    circuitos = leer_circuitos_excel(archivo_excel)
except FileNotFoundError:
    print(f"\n  ERROR: no se encontro '{archivo_excel}'")
    exit()

if len(circuitos) == 0:
    print("  ERROR: no se encontraron circuitos validos")
    exit()

# Generar reporte txt
lineas, total_ok, total_falla = generar_reporte_txt(
    nombre_proyecto, circuitos, fecha
)

print()
for linea in lineas:
    print(linea)

# Guardar txt
nombre_txt   = f"REPORTE_{nombre_proyecto.upper()}_{fecha_archivo}.txt"
nombre_excel = f"REPORTE_{nombre_proyecto.upper()}_{fecha_archivo}.xlsx"

guardar_txt(lineas, nombre_txt)
exportar_excel(nombre_proyecto, circuitos, fecha, nombre_excel)

print(f"\n  Proyecto  : {nombre_proyecto}")
print(f"  OK        : {total_ok}")
print(f"  FALLA     : {total_falla}")
print("\n  Listo.")