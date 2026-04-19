# ============================================================
# main.py
# Responsabilidad: flujo principal del programa
# Razón para cambiar: modificar interacción con el usuario
# ============================================================

import sys
import math
import os
import openpyxl
from datetime import datetime
from conductores import LIMITE_DV, TENSION_SISTEMA
from calculos import (
    capacidad_corregida, calcular_potencia,
    calcular_caida_tension, clasificar_caida, sugerir_conductor
)
from excel import (
    leer_circuitos_excel, leer_transformador_excel,
    leer_balance_excel, leer_tableros_excel,
    guardar_txt, exportar_excel, leer_perfil_excel,
    enriquecer_circuitos, leer_generador_excel, leer_sts_excel,
    leer_trafo_iso_excel, leer_ups_excel
)
from perfiles import obtener_perfil
from transformador import calcular_icc_transformador, icc_desde_tabla, clasificar_icc, reporte_transformador
from icc_punto import calcular_icc_todos_circuitos
from protecciones import verificar_circuito_completo, leer_protecciones_excel
from balance import calcular_balance_tableros, reporte_balance
from demanda import (
    calcular_demanda, seleccionar_transformador,
    dimensionar_acometida_sec, reporte_demanda
)
from excel import leer_demanda_excel, leer_cadena_excel
from coordinacion import verificar_cadena, reporte_coordinacion
from motores import calcular_motor
from generador import calcular_generador
from sts import calcular_sts
from trafo_iso import calcular_trafo_iso
from ups import calcular_ups

# ============================================================
# GENERACIÓN DE REPORTE TXT
# ============================================================

def generar_seccion_transformador(datos_trafo):
    lineas    = []
    resultado = {}

    if datos_trafo is None:
        lineas.append("  TRANSFORMADOR: no se encontró hoja 'Transformador'")
        lineas.append("  Icc en bornes BT: no calculada")
        return lineas, resultado

    modo = datos_trafo["modo"]

    if modo == "A":
        Icc_kA, Zt_ohm, info = calcular_icc_transformador(
            datos_trafo["kVA"], datos_trafo["Vn_BT"], datos_trafo["Ucc_pct"]
        )
        lineas_trafo = reporte_transformador(info, "A", Icc_kA)
    else:
        Icc_kA, Ucc_pct, kVA_ref = icc_desde_tabla(datos_trafo["kVA"])
        Zt_ohm = (Ucc_pct / 100) * (datos_trafo["Vn_BT"] ** 2 / (datos_trafo["kVA"] * 1000))
        info = {
            "kVA":     datos_trafo["kVA"],
            "Vn_BT":   datos_trafo["Vn_BT"],
            "Ucc_pct": Ucc_pct,
            "In_A":    round(datos_trafo["kVA"] * 1000 / (1.732 * datos_trafo["Vn_BT"]), 1),
            "Zt_ohm":  round(Zt_ohm, 6),
            "Icc_A":   round(Icc_kA * 1000, 1),
        }
        lineas_trafo = reporte_transformador(info, "B", Icc_kA)
        lineas_trafo.append(f"  Referencia tabla: {kVA_ref} kVA (IEC 60076)")

    lineas += lineas_trafo
    resultado = {
        "Icc_kA": Icc_kA, "Zt_ohm": Zt_ohm,
        "nivel": clasificar_icc(Icc_kA),
        "nombre": datos_trafo["nombre"], "modo": modo,
    }
    return lineas, resultado


def generar_seccion_motores(circuitos, perfil=None):
    perfil = perfil or {}
    lineas = []

    motores = [c for c in circuitos if str(c.get("tipo_carga", "")).lower() == "motor"]
    if not motores:
        return lineas

    lineas.append("")
    lineas.append("=" * 60)
    lineas.append("  MOTORES — CORRIENTE, ARRANQUE Y PROTECCIONES")
    lineas.append("  Normativa: RIC (Pliego Motores) / NCh Elec 4/2003")
    lineas.append("=" * 60)

    for m in motores:
        p_kw = m.get("P_kW")
        if p_kw is None:
            p_kw = round(calcular_potencia(m["I_diseno"], m["cos_phi"], m["sistema"]) / 1000.0, 2)

        resultado = calcular_motor(
            nombre=m.get("nombre"),
            P_kW=p_kw,
            V_nominal=TENSION_SISTEMA.get(m["sistema"], 380),
            cos_phi=m.get("cos_phi", 0.85),
            rendimiento=m.get("rendimiento", 0.92),
            sistema=m.get("sistema", "3F"),
            tipo_arranque=m.get("tipo_arranque", "directo"),
            regimen=m.get("regimen", "permanente"),
            periodo_min=m.get("periodo_min", 999),
            norma=perfil.get("norma", "AWG"),
        )
        guard = resultado["guardamotor"]
        prot = resultado["proteccion_arranque"]
        rango = resultado.get("arranque", {}).get("rango_tipico", (5.0, 8.0))
        i_n = float(resultado.get("I_n", resultado.get("I_plena_carga", 0.0)))
        arr_txt = {
            "directo": "arranque directo × 6",
            "estrella_triangulo": "arranque estrella-triangulo × 2",
            "variador": "arranque variador × 1",
        }.get(m.get("tipo_arranque", "directo"), "arranque directo × 6")

        lineas.append("")
        lineas.append(f"  Motor     : {resultado['nombre']}")
        lineas.append(f"  Sistema   : {resultado['sistema']} / {resultado['V_nominal']}V")
        lineas.append(
            f"  Potencia  : {resultado['P_kW']} kW  |  cos_phi: {resultado['cos_phi']}  |  η: {resultado['rendimiento']}"
        )
        lineas.append(f"  I_plena   : {resultado['I_plena_carga']} A")
        lineas.append(
            f"  I_arranque: {resultado['I_arranque']} A  ({arr_txt})"
            f"  — rango típico [{round(i_n * rango[0], 1)}-{round(i_n * rango[1], 1)}A]"
        )
        lineas.append(
            f"  Conductor : {resultado['conductor']} ({resultado['S_mm2']}mm2) — factor {resultado['factor_aplicado']} RIC"
        )
        lineas.append(
            f"  Guardamotor: {guard['rango_min']}-{guard['rango_max']}A  ajuste: {guard['ajuste_recomendado']}A"
        )
        lineas.append(
            f"  Protección: MA{int(round(guard['rango_max']))}A curva MA → "
            f"{'OK' if prot['ok'] else 'REVISAR'} (Im={prot['Im_min']}A > I_arr={resultado['I_arranque']}A)"
        )
    lineas.append("=" * 60)
    return lineas


def generar_seccion_generador(circuitos, datos_generador, protecciones=None, resultado_demanda=None):
    if not datos_generador:
        return []

    lineas = []
    p_demanda_kw = 0.0
    if resultado_demanda and resultado_demanda.get("P_total_kw") is not None:
        p_demanda_kw = float(resultado_demanda["P_total_kw"])
    else:
        p_demanda_kw = sum(
            calcular_potencia(c["I_diseno"], c["cos_phi"], c["sistema"]) / 1000.0
            for c in circuitos
        )

    motores = [c for c in circuitos if str(c.get("tipo_carga", "")).lower() == "motor"]
    if motores:
        motor_max = max(motores, key=lambda x: float(x.get("P_kW") or 0.0))
        p_motor_max = float(motor_max.get("P_kW") or 0.0)
        factor_por_tipo = {
            "directo": 6.0,
            "estrella_triangulo": 2.0,
            "variador": 1.2,
            "arranque_suave": 2.5,
        }
        factor_arr = factor_por_tipo.get(
            str(motor_max.get("tipo_arranque", "directo")).lower(),
            6.0,
        )
    else:
        p_motor_max = 0.0
        factor_arr = 1.0

    circuitos_prot = []
    if protecciones:
        for c in circuitos:
            p = protecciones.get(c.get("nombre"))
            if not p:
                continue
            circuitos_prot.append({
                "nombre": c.get("nombre"),
                "proteccion_A": p.get("In_A"),
                "curva": p.get("curva"),
            })

    resultado = calcular_generador(
        nombre=datos_generador["GE_nombre"],
        modelo_ge=datos_generador["GE_modelo"],
        P_ge_kVA_prime=datos_generador["GE_kVA_prime"],
        P_ge_kVA_emergencia=datos_generador["GE_kVA_emergencia"],
        cos_phi_ge=datos_generador.get("GE_cos_phi", 0.8),
        V_nominal=380.0,
        regimen_uso=datos_generador.get("GE_regimen_uso", "prime"),
        P_demanda_kW=p_demanda_kw,
        P_motor_max_kW=p_motor_max,
        factor_arranque_motor=factor_arr,
        altitud_msnm=datos_generador.get("GE_altitud_msnm", 0.0),
        Xd_pct=datos_generador.get("GE_Xd_pct", 25.0),
        consumo_100_galhr=datos_generador.get("GE_consumo_100_galhr"),
        consumo_75_galhr=datos_generador.get("GE_consumo_75_galhr"),
        capacidad_tanque_gal=datos_generador.get("GE_tanque_gal"),
        circuitos=circuitos_prot,
    )

    verif = resultado["verificacion_ge"]
    req = resultado["potencia_requerida"]
    icc = resultado["icc_ge"]
    dv = resultado["dv_arranque_ge"]
    autonomia = resultado.get("autonomia")

    margen_pct = (verif["margen_kVA"] / max(req["P_minimo_kVA"], 1e-9)) * 100.0
    ok_str = "OK" if verif["ok"] else "INSUFICIENTE"
    derrateo_pct = (1.0 - verif["factor_derrateo"]) * 100.0

    lineas.append("")
    lineas.append("=" * 60)
    lineas.append("  GENERADOR ELECTRICO - VERIFICACION Y CALCULOS")
    lineas.append("  Normativa: IEC 60034 / IEC 60909 / RIC")
    lineas.append("=" * 60)
    lineas.append(f"  GE            : {resultado['nombre']} ({resultado['modelo_ge']})")
    lineas.append(f"  Regimen uso   : {resultado['regimen_uso']}")
    lineas.append(
        f"  Altitud       : {round(resultado['altitud_msnm'], 1)} msnm -> derrateo: {round(derrateo_pct, 1)}% -> factor: {verif['factor_derrateo']:.3f}"
    )
    lineas.append("")
    lineas.append(
        f"  P disponible  : {round(resultado['P_ge_kVA_seleccionado'], 1)} kVA / {round(verif['P_ge_efectiva_kW'], 1)} kW ({resultado['regimen_uso']})"
    )
    lineas.append(
        f"  P requerida   : {round(req['P_minimo_kVA'], 0):.0f} kVA / {round(req['P_minimo_kW'], 0):.0f} kW"
    )
    lineas.append(f"  Margen        : {round(margen_pct, 1)}% -> {ok_str}")
    lineas.append("")
    lineas.append(f"  I_nominal GE  : {round(resultado['I_ge_nominal_A'], 1)} A (cos_phi={resultado['cos_phi_ge']})")
    lineas.append(f"  X'd asumido   : {round(icc['Xd_pct'], 1)}%")
    lineas.append(f"  Icc_GE nominal: {icc['Icc_nominal_kA']:.2f} kA")
    lineas.append(f"  Icc_GE max    : {icc['Icc_max_kA']:.2f} kA | min: {icc['Icc_min_kA']:.2f} kA")
    lineas.append("")
    lineas.append(
        f"  dV arranque motor mayor ({round(resultado['P_motor_max_kW'], 2)} kW DOL x {resultado['factor_arranque_motor']}):"
    )
    lineas.append(f"    dV = {round(dv['dv_pct'], 1)}% -> {dv['estado']}")

    if autonomia:
        lineas.append("")
        lineas.append("  Autonomia estimada:")
        lineas.append(
            f"    Uso: {round(autonomia['uso_pct'], 1)}% -> consumo ~{round(autonomia['consumo_estimado_galhr'], 1)} gal/hr"
        )
        lineas.append(
            f"    Tanque {round(float(datos_generador.get('GE_tanque_gal') or 0), 1)} gal -> autonomia estimada: {round(autonomia['autonomia_hr'], 1)} hr -> {'OK' if autonomia['autonomia_ok'] else 'REVISAR'}"
        )

    lineas.append("")
    lineas.append("  Verificacion protecciones modo GE:")
    if resultado["protecciones_modo_ge"]:
        for p in resultado["protecciones_modo_ge"]:
            lineas.append(
                f"    {p['nombre']}: {p['proteccion']} | Im={round(p['Im'], 1)}A vs Icc_GE={round(p['Icc_ge_A'], 1)}A -> {p['observacion']}"
            )
    else:
        lineas.append("    Sin datos de protecciones para verificar.")
    lineas.append("=" * 60)
    return lineas


def generar_seccion_sts(datos_sts):
    if not datos_sts:
        return []

    r = calcular_sts(
        nombre=datos_sts["STS_nombre"],
        modelo_sts=datos_sts["STS_modelo"],
        P_modulo_kVA=datos_sts["STS_P_modulo_kVA"],
        n_modulos=datos_sts.get("STS_n_modulos", 1),
        t_transferencia_ms=datos_sts["STS_t_transferencia_ms"],
        V_nominal=datos_sts.get("STS_V_nominal", 380.0),
        P_carga_kVA=datos_sts["STS_P_carga_kVA"],
        cos_phi_carga=datos_sts.get("STS_cos_phi", 0.9),
        tipo_carga=datos_sts.get("STS_tipo_carga", "general"),
        topologia=datos_sts.get("STS_topologia", "simple"),
        n_sts=datos_sts.get("STS_n_sts", 1),
        P_no_lineal_kVA=datos_sts.get("STS_P_no_lineal_kVA", 0.0),
        t_sobrecarga_seg=datos_sts.get("STS_t_sobrecarga_seg", 0.0),
    )

    cap = r["capacidad"]
    tr = r["transferencia"]
    ov = r["overload"]
    red = r.get("redundancia_2N")
    nl = r.get("carga_no_lineal")
    topo_txt = "2N" if r["topologia"] == "2n" else r["topologia"]

    lineas = []
    lineas.append("")
    lineas.append("=" * 60)
    lineas.append("  STS - TRANSFERENCIA ESTATICA")
    lineas.append("  Normativa: IEC 60947-6-1 / IEC 62040-3 / IEC 60364-5-55")
    lineas.append("=" * 60)
    lineas.append(f"  STS           : {r['nombre']} ({r['modelo_sts']})")
    lineas.append(f"  Topologia     : {topo_txt}")
    lineas.append(
        f"  Modulos       : {r['n_modulos']} x {r['P_modulo_kVA']}kVA = {r['P_sts_total_kVA']}kVA por STS"
    )
    lineas.append(f"  Tension       : {r['V_nominal']}V 3F / 50Hz")
    lineas.append("")
    lineas.append("  CAPACIDAD")
    lineas.append(f"  P_carga total : {r['P_carga_kVA']}kVA")
    lineas.append(f"  Uso           : {cap['uso_pct']}% -> {'OK' if cap['ok'] else 'REVISAR'}")
    lineas.append(f"  Margen        : {cap['margen_kVA']}kVA")

    if r["topologia"] == "2n" and red:
        lineas.append("")
        lineas.append("  REDUNDANCIA 2N")
        lineas.append(
            f"  Uso normal    : {red['uso_normal_pct']}% por STS -> {'OK' if red['ok_normal'] else 'REVISAR'}"
        )
        lineas.append(
            f"  Uso falla BUS : {red['uso_falla_pct']}% en STS sobreviviente -> {'OK' if red['ok_falla'] else 'REVISAR'}"
        )

    lineas.append("")
    lineas.append("  TRANSFERENCIA")
    lineas.append(f"  t_transfer    : {tr['t_transferencia_ms']} ms")
    lineas.append(f"  Tipo carga    : {r['tipo_carga']}")
    lineas.append(f"  t_max         : {tr['t_max_ms']}ms ({tr['norma']})")
    lineas.append(f"  Margen        : {tr['margen_ms']}ms -> {'OK' if tr['ok'] else 'REVISAR'}")

    if nl:
        lineas.append("")
        lineas.append("  CARGA NO LINEAL")
        lineas.append(
            f"  P_no_lineal   : {r['P_no_lineal_kVA']}kVA / {r['P_carga_kVA']}kVA -> {nl['pct_no_lineal']}%"
        )
        lineas.append(f"  Factor cresta : >= {nl['factor_cresta_requerido']}:1 requerido")
        lineas.append(f"  Estado        : {'OK' if nl['ok'] else 'ALERTAR'}")

    if r["t_sobrecarga_seg"] > 0:
        lineas.append("")
        lineas.append("  OVERLOAD")
        lineas.append(f"  Nivel         : {ov['nivel']} ({ov['sobrecarga_pct']}%)")
        lineas.append(
            f"  t_max permit  : {ov['t_max_permitido_seg']}s -> {'OK' if ov['ok'] else 'REVISAR'}"
        )

    lineas.append("=" * 60)
    return lineas


def generar_seccion_trafo_iso(datos_trafo_iso):
    if not datos_trafo_iso:
        return []

    r = calcular_trafo_iso(
        nombre=datos_trafo_iso["TISO_nombre"],
        P_trafo_kVA=datos_trafo_iso["TISO_P_kVA"],
        V_primario=datos_trafo_iso["TISO_V_primario"],
        V_secundario=datos_trafo_iso["TISO_V_secundario"],
        conexion=datos_trafo_iso.get("TISO_conexion", "Dyn5"),
        P_carga_kVA=datos_trafo_iso["TISO_P_carga_kVA"],
        cos_phi=datos_trafo_iso.get("TISO_cos_phi", 0.9),
        Ucc_pct=datos_trafo_iso.get("TISO_Ucc_pct", 5.0),
        n_trafos=datos_trafo_iso.get("TISO_n_trafos", 1),
        modo=datos_trafo_iso.get("TISO_modo", "servicio"),
    )
    cap = r["capacidad"]
    icc = r["icc_secundario"]
    dv = r["dv_trafo"]

    lineas = []
    lineas.append("")
    lineas.append("=" * 60)
    lineas.append("  TRANSFORMADOR DE AISLAMIENTO")
    lineas.append("  Normativa: IEC 60076-1 / IEC 61558-2-4 / RIC N10")
    lineas.append("=" * 60)
    lineas.append(f"  Trafo         : {r['nombre']} ({r['conexion']})")
    lineas.append(f"  Configuracion : {r['n_trafos']} unidad(es) - {r['modo']}")
    lineas.append(f"  P nominal     : {r['P_total_kVA']}kVA | V: {r['V_primario']}/{r['V_secundario']}V")
    lineas.append(f"  P carga       : {r['P_carga_kVA']}kVA")
    lineas.append(f"  Uso           : {cap['uso_pct']}% -> {'OK' if cap['ok'] else 'REVISAR'}")
    lineas.append(f"  I_nominal sec : {r['I_nominal_sec_A']}A")
    lineas.append(f"  Ucc%          : {r['Ucc_pct']}% (tol. +-7.5%)")
    lineas.append(
        f"  Icc sec nom   : {icc['Icc_nominal_kA']}kA | max: {icc['Icc_max_kA']}kA | min: {icc['Icc_min_kA']}kA"
    )
    lineas.append(f"  dV trafo      : {dv['dv_pct']}% -> {'OK' if dv['ok'] else 'REVISAR'}")
    lineas.append("=" * 60)
    return lineas


def generar_seccion_ups(datos_ups):
    if not datos_ups:
        return []

    r = calcular_ups(
        nombre=datos_ups["UPS_nombre"],
        modelo_ups=datos_ups["UPS_modelo"],
        tipo_ups=datos_ups.get("UPS_tipo", "VFI"),
        P_ups_kVA=datos_ups["UPS_P_kVA"],
        V_nominal=datos_ups.get("UPS_V_nominal", 380.0),
        P_carga_kW=datos_ups["UPS_P_carga_kW"],
        cos_phi_carga=datos_ups.get("UPS_cos_phi", 0.9),
        tipo_carga=datos_ups.get("UPS_tipo_carga", "general"),
        nivel_infraestructura=datos_ups.get("UPS_nivel_infraestructura", "critico"),
        n_baterias_serie=datos_ups["UPS_n_baterias_serie"],
        V_bat_unitaria=datos_ups["UPS_V_bat"],
        Ah_bat=datos_ups["UPS_Ah_bat"],
        n_strings=datos_ups["UPS_n_strings"],
        temperatura=datos_ups.get("UPS_temperatura", 25.0),
    )

    cap = r["capacidad"]
    bat = r["banco_baterias"]
    aut = r["autonomia"]
    rec = r["recarga"]
    tv = r["tipo_validacion"]

    lineas = []
    lineas.append("")
    lineas.append("=" * 60)
    lineas.append("  UPS - SISTEMA DE ALIMENTACION ININTERRUMPIDA")
    lineas.append("  Normativa: IEC 62040-1/2/3/4 / TIA-942")
    lineas.append("=" * 60)
    lineas.append(f"  UPS           : {r['nombre']} ({r['modelo_ups']})")
    lineas.append(f"  Tipo IEC      : {r['tipo_ups']} - {tv['tipo_descripcion']}")
    lineas.append(f"  P nominal     : {r['P_ups_kVA']}kVA | V: {r['V_nominal']}V")
    lineas.append(f"  P carga       : {r['P_carga_kW']}kW ({cap['uso_pct']}%) -> {'OK' if cap['ok'] else 'REVISAR'}")
    lineas.append("")
    lineas.append("  BANCO BATERIAS")
    lineas.append(
        f"  Config        : {r['n_strings']} strings x {r['n_baterias_serie']} baterias x {r['V_bat_unitaria']}V/{r['Ah_bat']}Ah"
    )
    lineas.append(f"  V string      : {bat['V_string']}V DC")
    lineas.append(f"  Ah efectivos  : {bat['Ah_efectivo']}Ah (temp {r['temperatura']}C, factor {bat['factor_temp']})")
    lineas.append(f"  Energia total : {bat['E_kWh']}kWh")
    lineas.append("")
    lineas.append("  AUTONOMIA")
    lineas.append(f"  P en baterias : {aut['P_baterias_kW']}kW (eta_ups={r['eta_ups']})")
    lineas.append(f"  Autonomia     : {aut['t_min']} min -> {aut['estado']}")
    lineas.append(f"  Minimo normado: {aut['t_minimo_normado']} min ({aut['norma_aplicada']})")
    lineas.append("")
    lineas.append("  RECARGA")
    lineas.append(f"  I recarga     : {rec['I_carga_A']}A")
    lineas.append(
        f"  Tiempo recarga: {rec['t_recarga_hr']}hr -> {'OK' if rec['ok'] else 'REVISAR'} (IEC 62040-4: <= 12hr)"
    )
    lineas.append("=" * 60)
    return lineas


def generar_reporte_txt(nombre_proyecto, circuitos, fecha,
                        datos_trafo=None, protecciones=None,
                        balance_datos=None, tableros_datos=None,
                        params_demanda=None,
                        cadena_datos=None, perfil=None,
                        datos_generador=None, datos_sts=None,
                        datos_trafo_iso=None, datos_ups=None):
    perfil = perfil or {}
    lineas      = []
    total_ok    = 0
    total_falla = 0
    prot_ok     = 0
    prot_falla  = 0

    # --- ENCABEZADO ---
    lineas.append("=" * 60)
    lineas.append(f"  REPORTE — {nombre_proyecto}")
    lineas.append(f"  Fecha        : {fecha}")
    lineas.append(f"  Normativa    : SEC RIC N10 / NEC / IEC 60364")
    lineas.append(f"  Limite caida : {LIMITE_DV}% circuito final / 5% total")
    lineas.append(f"  Circuitos    : {len(circuitos)}")
    lineas.append("=" * 60)

    # --- TRANSFORMADOR ---
    lineas.append("")
    lineas_trafo, resultado_trafo = generar_seccion_transformador(datos_trafo)
    lineas += lineas_trafo

    # --- Icc POR PUNTO — M2 ---
    if resultado_trafo and "Zt_ohm" in resultado_trafo:
        circuitos = calcular_icc_todos_circuitos(
            resultado_trafo["Zt_ohm"], circuitos
        )

    # --- CIRCUITOS ---
    lineas.append("")
    lineas.append("=" * 60)
    lineas.append("  CIRCUITOS — CAÍDA DE TENSIÓN, Icc Y PROTECCIONES")
    lineas.append("=" * 60)

    for c in circuitos:
        I_cap        = capacidad_corregida(c["I_max"], c["paralelos"], c["temp_amb"])
        P_watts      = calcular_potencia(c["I_diseno"], c["cos_phi"], c["sistema"])
        dV_V, dV_pct = calcular_caida_tension(
            c["L_m"], c["S_mm2"], c["I_diseno"], c["paralelos"], c["sistema"]
        )
        estado_dV = clasificar_caida(dV_pct)
        estado_I  = "OK" if c["I_diseno"] <= I_cap else "SUPERA"
        desc_cond = (
            f"{c['paralelos']}x{c['conductor']} (S={c['S_mm2']*c['paralelos']}mm2)"
            if c["paralelos"] > 1
            else f"{c['conductor']} ({c['S_mm2']}mm2)"
        )
        V_nom = TENSION_SISTEMA[c["sistema"]]

        lineas.append("")
        lineas.append(f"  Circuito  : {c['nombre']}")
        lineas.append(f"  Sistema   : {c['sistema']} / {V_nom}V | Temp: {c['temp_amb']}C")
        lineas.append(f"  Conductor : {desc_cond}")
        lineas.append(f"  Corriente : {c['I_diseno']}A -> {estado_I} (cap. {I_cap}A)")
        lineas.append(f"  Potencia  : {P_watts} W")
        lineas.append(f"  Caida dV  : {dV_V}V ({dV_pct}%) -> {estado_dV}")

        if "Icc_kA" in c:
            lineas.append(f"  Icc punto : {c['Icc_kA']} kA  -> {c['nivel_icc']}")

        if protecciones and c["nombre"] in protecciones:
            p = protecciones[c["nombre"]]
            r = verificar_circuito_completo(
                c["nombre"], p["In_A"], p["curva"],
                p["poder_corte_kA"], c.get("Icc_kA", 0), V_nom
            )
            lineas.append(
                f"  Proteccion: {p['curva']}{int(p['In_A'])}A / "
                f"{int(p['poder_corte_kA'])}kA -> {r['estado']}"
            )
            lineas.append(
                f"  Im_min    : {r['Im_min_A']}A | "
                f"margen {r['margen_pct']}% -> {r['clasif_margen']}"
            )
            if r["estado"] == "OK":
                prot_ok += 1
            else:
                prot_falla += 1

        if estado_dV == "FALLA" or estado_I == "SUPERA":
            cond, mm2, dv = sugerir_conductor(
                c["L_m"], c["I_diseno"], c["paralelos"],
                c["sistema"], c["temp_amb"],
                norma=perfil.get("norma", "AWG")
            )
            lineas.append(
                f"  SUGERENCIA: usar {cond} ({mm2}mm2) -> dV={dv}%"
                if cond else
                "  SUGERENCIA: ningun conductor en tabla es suficiente"
            )
            total_falla += 1
        else:
            total_ok += 1

    # --- BALANCE DE CARGA — M4 ---
    resultado_balance = None
    if balance_datos and tableros_datos:
        kVA_trafo = datos_trafo["kVA"] if datos_trafo else 1000
        resultado_balance = calcular_balance_tableros(
            circuitos, balance_datos, tableros_datos, kVA_trafo
        )
        lineas.append("")
        lineas += reporte_balance(resultado_balance)

    # --- DEMANDA MÁXIMA — M6 ---
    resultado_demanda = None
    if params_demanda and balance_datos:
        resultado_demanda = calcular_demanda(
            circuitos, balance_datos, params_demanda
        )
        resultado_trafo_m6 = None
        resultado_sec_m6   = None

        if params_demanda.get("tipo_alimentador") == "transformador":
            resultado_trafo_m6 = seleccionar_transformador(
                resultado_demanda["S_futuro_kva"]
            )
        else:
            resultado_sec_m6 = dimensionar_acometida_sec(
                resultado_demanda["S_futuro_kva"],
                params_demanda["tension_alim"],
                params_demanda["sistema_alim"],
                params_demanda.get("zona_sec", "urbana")
            )

        lineas.append("")
        lineas += reporte_demanda(
            resultado_demanda, resultado_trafo_m6, resultado_sec_m6
        )

    # --- COORDINACIÓN TCC — M7 ---
    if cadena_datos:
        # Agrupar dispositivos por modo
        modos = {}
        for d in cadena_datos:
            m = d.get("modo", "red")
            modos.setdefault(m, []).append(d)

        for modo, dispositivos in modos.items():
            # Ordenar por nivel
            dispositivos_ord = sorted(dispositivos, key=lambda x: x["nivel"])
            # Usar Icc del primer dispositivo disponible, o 0
            Icc_A = 0
            for d in reversed(dispositivos_ord):
                if d.get("Icc_kA"):
                    Icc_A = d["Icc_kA"] * 1000
                    break
            if Icc_A > 0:
                res_cadena = verificar_cadena(
                    dispositivos_ord, Icc_A, sistema="3F_380"
                )
                lineas.append("")
                lineas += reporte_coordinacion(
                    res_cadena, f"Modo {modo}"
                )

    # --- MOTORES --- M8
    lineas += generar_seccion_motores(circuitos, perfil=perfil)

    # --- GENERADOR --- M9
    lineas += generar_seccion_generador(
        circuitos,
        datos_generador,
        protecciones=protecciones,
        resultado_demanda=resultado_demanda,
    )

    # --- STS --- M11
    lineas += generar_seccion_sts(datos_sts)

    # --- TRAFO ISO --- M12A
    lineas += generar_seccion_trafo_iso(datos_trafo_iso)

    # --- UPS --- M12B
    lineas += generar_seccion_ups(datos_ups)

    # --- RESUMEN FINAL ---
    lineas.append("")
    lineas.append("=" * 60)
    lineas.append(f"  Circuitos OK    : {total_ok}")
    lineas.append(f"  Circuitos FALLA : {total_falla}")
    if resultado_trafo:
        lineas.append(f"  Icc bornes BT   : {resultado_trafo['Icc_kA']} kA")
        lineas.append(f"  Nivel Icc       : {resultado_trafo['nivel']}")
    if protecciones:
        lineas.append(f"  Protecciones OK : {prot_ok}")
        lineas.append(f"  Protec. FALLA   : {prot_falla}")
    if resultado_balance:
        lineas.append(f"  Uso transf.     : {resultado_balance['uso_trafo_pct']}%"
                      f" -> {resultado_balance['estado_trafo']}")
    lineas.append("=" * 60)

    return lineas, total_ok, total_falla

# ============================================================
# PROGRAMA PRINCIPAL
# ============================================================

print("=" * 60)
print("  MOTOR DE CALCULO BT — VERSION MODULAR")
print("  ΔV | Icc | Protecciones | Balance de carga")
print("  Normativa: SEC RIC N10 / NEC / IEC 60364")
print("=" * 60)

ahora         = datetime.now()
fecha         = ahora.strftime("%d/%m/%Y %H:%M")
fecha_archivo = ahora.strftime("%Y%m%d_%H%M")

nombre_proyecto = input("\n  Nombre del proyecto : ").strip()
archivo_excel   = input("  Archivo Excel       : ").strip()

if not archivo_excel.endswith(".xlsx"):
    archivo_excel += ".xlsx"

# --- LEER TODAS LAS HOJAS ---
datos_trafo        = leer_transformador_excel(archivo_excel)
protecciones_excel = {}
balance_datos      = {}
tableros_datos     = {}
params_demanda     = None
cadena_datos       = []
datos_generador    = None
datos_sts          = None
datos_trafo_iso    = None
datos_ups          = None
perfil             = obtener_perfil("industrial").copy()

try:
    _libro = openpyxl.load_workbook(archivo_excel, data_only=True)
    datos_perfil = leer_perfil_excel(_libro) or {}
    perfil_clave = datos_perfil.get("perfil", "industrial")
    perfil = obtener_perfil(perfil_clave).copy()

    hoja_perfil = next(
        (_libro[nombre] for nombre in _libro.sheetnames if nombre.lower() == "perfil"),
        None
    )
    norma_explicitada = False
    if hoja_perfil:
        for fila in hoja_perfil.iter_rows(min_row=2, values_only=True):
            if fila[0] and str(fila[0]).strip().lower() == "norma":
                norma_explicitada = True
                break
    if norma_explicitada:
        perfil["norma"] = datos_perfil.get("norma", "AWG").upper()

    protecciones_excel = leer_protecciones_excel(_libro)
    balance_datos      = leer_balance_excel(_libro)
    tableros_datos     = leer_tableros_excel(_libro)

    if protecciones_excel:
        print(f"  Protecciones  : {len(protecciones_excel)} circuitos")
    if balance_datos:
        print(f"  Balance       : {len(balance_datos)} circuitos / "
              f"{len(tableros_datos)} tableros")
    params_demanda = leer_demanda_excel(_libro)
    if params_demanda:
        print(f"  Demanda M6    : {params_demanda['tipo_instalacion']} / "
              f"{params_demanda['tipo_alimentador']}")
    cadena_datos = leer_cadena_excel(_libro)
    datos_generador = leer_generador_excel(_libro)
    datos_sts = leer_sts_excel(_libro)
    datos_trafo_iso = leer_trafo_iso_excel(_libro)
    datos_ups = leer_ups_excel(_libro)
    if datos_generador:
        print(
            f"  Generador M9  : {datos_generador['GE_nombre']} ({datos_generador['GE_modelo']})"
        )
    if datos_sts:
        print(
            f"  STS M11       : {datos_sts['STS_nombre']} ({datos_sts['STS_modelo']})"
        )
    if datos_trafo_iso:
        print(
            f"  Trafo ISO M12 : {datos_trafo_iso['TISO_nombre']} ({datos_trafo_iso['TISO_conexion']})"
        )
    if datos_ups:
        print(
            f"  UPS M12       : {datos_ups['UPS_nombre']} ({datos_ups['UPS_modelo']})"
        )
    if cadena_datos:
        print(f"  Coordinación  : {len(cadena_datos)} dispositivos en cadena")
except Exception as e:
    print(f"  AVISO lectura hojas opcionales: {e}")
    params_demanda = None
    cadena_datos   = []
    datos_generador = None
    datos_sts = None
    datos_trafo_iso = None
    datos_ups = None

# --- LEER CIRCUITOS ---
try:
    circuitos = leer_circuitos_excel(archivo_excel)
    circuitos = enriquecer_circuitos(
        circuitos, norma=perfil.get("norma", "AWG")
    )
except FileNotFoundError as e:
    print(f"\n  ERROR: {e}")
    input("\n  Presiona Enter para cerrar...")
    sys.exit()
except ValueError as e:
    print(f"\n  ERROR: {e}")
    input("\n  Presiona Enter para cerrar...")
    sys.exit()
except PermissionError as e:
    print(f"\n  ERROR: {e}")
    input("\n  Presiona Enter para cerrar...")
    sys.exit()
except Exception as e:
    print(f"\n  ERROR inesperado: {e}")
    print("  Contacta al desarrollador con este mensaje.")
    input("\n  Presiona Enter para cerrar...")
    sys.exit()

if len(circuitos) == 0:
    print("\n  ERROR: no se procesó ningún circuito válido.")
    input("\n  Presiona Enter para cerrar...")
    sys.exit()

# --- GENERAR Y MOSTRAR REPORTE ---
lineas, total_ok, total_falla = generar_reporte_txt(
    nombre_proyecto, circuitos, fecha,
    datos_trafo, protecciones_excel,
    balance_datos, tableros_datos,
    params_demanda, cadena_datos,
    perfil=perfil, datos_generador=datos_generador, datos_sts=datos_sts,
    datos_trafo_iso=datos_trafo_iso, datos_ups=datos_ups
)

print()
for linea in lineas:
    print(linea)

# --- GUARDAR ---
nombre_txt  = f"REPORTE_{nombre_proyecto.upper()}_{fecha_archivo}.txt"
nombre_xlsx = f"REPORTE_{nombre_proyecto.upper()}_{fecha_archivo}.xlsx"

guardar_txt(lineas, nombre_txt)
exportar_excel(nombre_proyecto, circuitos, fecha, nombre_xlsx, perfil=perfil)

try:
    from persistencia import registrar_ejecucion

    max_dv_pct = 0.0
    max_icc_ka = 0.0
    circuitos_persistencia = []
    icc_por_circuito = {}
    try:
        if datos_trafo:
            if datos_trafo.get("modo") == "A":
                _, zt_ohm, _ = calcular_icc_transformador(
                    datos_trafo["kVA"], datos_trafo["Vn_BT"], datos_trafo["Ucc_pct"]
                )
            else:
                _, ucc_pct_ref, _ = icc_desde_tabla(datos_trafo["kVA"])
                zt_ohm = (ucc_pct_ref / 100.0) * (
                    (datos_trafo["Vn_BT"] ** 2) / (datos_trafo["kVA"] * 1000.0)
                )
            circuitos_icc = calcular_icc_todos_circuitos(zt_ohm, circuitos)
            for c_icc in circuitos_icc:
                icc_por_circuito[c_icc.get("nombre")] = c_icc.get("Icc_kA")
    except Exception:
        icc_por_circuito = {}

    for c in circuitos:
        dV_V, dV_pct = calcular_caida_tension(
            c["L_m"], c["S_mm2"], c["I_diseno"], c["paralelos"], c["sistema"]
        )
        I_cap = capacidad_corregida(c["I_max"], c["paralelos"], c["temp_amb"])
        estado_dV = clasificar_caida(dV_pct)
        estado_I = "OK" if c["I_diseno"] <= I_cap else "SUPERA"
        estado = "OK" if (estado_dV != "FALLA" and estado_I == "OK") else "CON_FALLAS"
        icc_ka = c.get("icc_ka") or c.get("Icc_kA") or icc_por_circuito.get(c.get("nombre"))

        observaciones = c.get("nivel_icc")
        if estado_dV == "FALLA" or estado_I == "SUPERA":
            cond, mm2, dv = sugerir_conductor(
                c["L_m"], c["I_diseno"], c["paralelos"],
                c["sistema"], c["temp_amb"],
                norma=perfil.get("norma", "AWG")
            )
            if cond:
                observaciones = f"redimensionar a {cond} ({mm2}mm2), dV={dv}%"
            else:
                observaciones = "redimensionar conductor (sin alternativa en tabla)"

        circuitos_persistencia.append(
            {
                "nombre": c.get("nombre"),
                "conductor": c.get("conductor"),
                "norma": perfil.get("norma", "AWG"),
                "S_mm2": c.get("S_mm2"),
                "I_diseno": c.get("I_diseno"),
                "I_max": c.get("I_max"),
                "cos_phi": c.get("cos_phi"),
                "L_m": c.get("L_m"),
                "paralelos": c.get("paralelos"),
                "sistema": c.get("sistema"),
                "dv_v": round(dV_V, 3),
                "dv_pct": round(dV_pct, 3),
                "icc_ka": icc_ka,
                "estado": estado,
                "observaciones": observaciones,
            }
        )
        if dV_pct > max_dv_pct:
            max_dv_pct = dV_pct
        if (icc_ka or 0.0) > max_icc_ka:
            max_icc_ka = icc_ka or 0.0

    datos_transformador = None
    try:
        if datos_trafo:
            kVA = float(datos_trafo["kVA"])
            vn_bt = float(datos_trafo["Vn_BT"])
            if datos_trafo.get("modo") == "A":
                icc_nom_kA, _, _ = calcular_icc_transformador(
                    datos_trafo["kVA"], datos_trafo["Vn_BT"], datos_trafo["Ucc_pct"]
                )
                ucc_pct = float(datos_trafo["Ucc_pct"])
            else:
                icc_nom_kA, ucc_pct, _ = icc_desde_tabla(datos_trafo["kVA"])

            z_min = ((ucc_pct * 0.925) / 100.0) * (vn_bt ** 2 / (kVA * 1000.0))
            z_max = ((ucc_pct * 1.075) / 100.0) * (vn_bt ** 2 / (kVA * 1000.0))
            icc_max_kA = round((1.1 * vn_bt / (1.732 * z_min)) / 1000.0, 2) if z_min > 0 else None
            icc_min_kA = round((0.95 * vn_bt / (1.732 * z_max)) / 1000.0, 2) if z_max > 0 else None

            datos_transformador = {
                "kVA": round(kVA, 2),
                "Vn_BT": round(vn_bt, 2),
                "Ucc_pct": round(ucc_pct, 2),
                "Icc_nom_kA": round(icc_nom_kA, 2),
                "Icc_max_kA": icc_max_kA,
                "Icc_min_kA": icc_min_kA,
            }
    except Exception:
        datos_transformador = None

    datos_balance_demanda = {}
    try:
        if balance_datos and tableros_datos:
            kVA_trafo = datos_trafo["kVA"] if datos_trafo else 1000
            r_balance = calcular_balance_tableros(
                circuitos, balance_datos, tableros_datos, kVA_trafo
            )
            datos_balance_demanda["balance"] = {
                "S_total_kva": r_balance.get("S_total_kva"),
                "uso_trafo_pct": r_balance.get("uso_trafo_pct"),
                "kVA_trafo": r_balance.get("kVA_trafo"),
            }
        if params_demanda and balance_datos:
            r_dem = calcular_demanda(circuitos, balance_datos, params_demanda)
            datos_balance_demanda["demanda"] = {
                "P_total_kw": r_dem.get("P_total_kw"),
                "S_total_kva": r_dem.get("S_total_kva"),
                "factor_crecimiento": r_dem.get("factor_crecimiento"),
                "S_futuro_kva": r_dem.get("S_futuro_kva"),
            }
    except Exception:
        pass

    datos_run = {
        "project_id": nombre_proyecto,
        "revision": "CLI",
        "timestamp": datetime.now().astimezone().isoformat(),
        "perfil": perfil.get("label", "industrial"),
        "norma": perfil.get("norma", "AWG"),
        "n_circuitos": len(circuitos),
        "n_ok": total_ok,
        "n_advertencias": max(len(circuitos) - total_ok - total_falla, 0),
        "n_fallas": total_falla,
        "max_dv_pct": round(max_dv_pct, 3),
        "max_icc_ka": round(max_icc_ka, 3),
        "status": "OK" if total_falla == 0 else "CON_FALLAS",
        "ruta_reporte_txt": nombre_txt,
        "ruta_reporte_xlsx": nombre_xlsx,
        "circuitos": circuitos_persistencia,
        "transformador": datos_transformador,
        "balance_demanda": datos_balance_demanda,
    }

    # Reporteria SEC adicional (sin interrumpir flujo principal).
    try:
        from reporteria_sec import generar_memoria_docx, generar_reporte_pdf

        carpeta_salida = os.getcwd()
        ruta_docx = generar_memoria_docx(datos_run, circuitos_persistencia, carpeta_salida)
        ruta_pdf = generar_reporte_pdf(datos_run, circuitos_persistencia, carpeta_salida)
        datos_run["ruta_reporte_docx"] = ruta_docx
        datos_run["ruta_reporte_pdf"] = ruta_pdf
        print(f"  Memoria SEC : {ruta_docx}")
        print(f"  Reporte PDF : {ruta_pdf}")
    except Exception as e:
        print(f"  Advertencia reporteria: {e}")

    registrar_ejecucion(datos_run)
except Exception as e:
    print(f"  Aviso persistencia: {e}")

print(f"\n  Proyecto  : {nombre_proyecto}")
print(f"  OK        : {total_ok}")
print(f"  FALLA     : {total_falla}")
print("\n  Listo.")
input("\n  Presiona Enter para cerrar...")
