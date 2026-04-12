# ============================================================
# excel.py
# Responsabilidad: lectura de entrada y exportación de reportes
# Razón para cambiar: modificar formato o estructura de archivos
# ============================================================

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.exceptions import InvalidFileException
from conductores import (
    CONDUCTORES, get_tabla_conductores,
    FACTOR_SISTEMA, TENSION_SISTEMA, LIMITE_DV, FACTORES_TEMP
)
from calculos import (
    capacidad_corregida, calcular_potencia,
    calcular_caida_tension, clasificar_caida, sugerir_conductor
)

# ============================================================
# LECTURA — TRANSFORMADOR
# ============================================================

def leer_transformador_excel(nombre_archivo):
    """
    Lee datos del transformador desde la hoja 'Transformador'.
    Estructura: columna A = campo, columna B = valor
    Retorna diccionario o None si no existe la hoja.
    """
    try:
        libro = openpyxl.load_workbook(nombre_archivo, data_only=True)
    except Exception:
        return None

    # Buscar hoja — insensible a mayúsculas
    hoja_nombre = None
    for nombre in libro.sheetnames:
        if nombre.lower() == "transformador":
            hoja_nombre = nombre
            break

    if not hoja_nombre:
        print("  INFO: no se encontró hoja 'Transformador' — omitiendo cálculo de Icc")
        return None

    hoja   = libro[hoja_nombre]
    datos  = {}
    errores = []

    # Leer pares campo:valor desde fila 2
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        campo = fila[0]
        valor = fila[1]
        if campo is None:
            continue
        # Normalizar campo — minúsculas sin tildes
        campo_norm = str(campo).strip().lower()
        for orig, repl in [("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u"),("ñ","n")]:
            campo_norm = campo_norm.replace(orig, repl)
        datos[campo_norm] = valor

    modo = str(datos.get("modo", "B")).strip().upper()

    # Modo A requiere datos de placa completos
    if modo == "A":
        for campo in ["kva", "vn_bt", "ucc_pct"]:
            if campo not in datos or datos[campo] is None:
                errores.append(f"Modo A requiere campo '{campo}'")

    if errores:
        print("\n  ADVERTENCIAS transformador:")
        for e in errores:
            print(f"  ⚠ {e}")
        return None

    return {
        "nombre":   str(datos.get("nombre", "Transformador")).strip(),
        "modo":     modo,
        "kVA":      float(datos.get("kva", 0)),
        "Vn_BT":    float(datos.get("vn_bt", 380)),
        "Ucc_pct":  float(datos.get("ucc_pct", 5.0)) if modo == "A" else None,
        "conexion": str(datos.get("conexion", datos.get("conexion", "Dyn11"))).strip(),
    }

# ============================================================
# LECTURA — CIRCUITOS
# ============================================================

def leer_circuitos_excel(nombre_archivo):
    """
    Lee circuitos desde la hoja 'circuitos' del Excel.
    Fila 1 = encabezados. Datos desde fila 2.
    Columnas: nombre | sistema | conductor | paralelos |
              I_diseno | cos_phi | L_m | temp_amb
    Retorna lista de diccionarios.
    """
    circuitos = []
    errores   = []

    # --- NIVEL 1: verificar que el archivo existe y es válido ---
    try:
        libro = openpyxl.load_workbook(nombre_archivo, data_only=True)
    except FileNotFoundError:
        raise FileNotFoundError(
            f"No se encontró '{nombre_archivo}'.\n"
            f"Verifica que el archivo está en la misma carpeta que el programa."
        )
    except InvalidFileException:
        raise ValueError(
            f"'{nombre_archivo}' no es un archivo Excel válido.\n"
            f"Verifica que tiene extensión .xlsx y no está dañado."
        )
    except PermissionError:
        raise PermissionError(
            f"No se puede leer '{nombre_archivo}'.\n"
            f"Cierra el archivo en Excel antes de ejecutar el programa."
        )

    # Buscar hoja circuitos — insensible a mayúsculas
    hoja_nombre = None
    for nombre in libro.sheetnames:
        if nombre.lower() == "circuitos":
            hoja_nombre = nombre
            break

    if not hoja_nombre:
        # Si no encuentra hoja 'circuitos' usa la primera hoja
        hoja_nombre = libro.sheetnames[0]

    hoja = libro[hoja_nombre]

    print(f"\n  Leyendo: {nombre_archivo}")
    print(f"  Circuitos encontrados: {hoja.max_row - 1}")

    # --- NIVEL 2: verificar cada fila de datos ---
    for num_fila, fila in enumerate(
        hoja.iter_rows(min_row=2, values_only=True), start=2
    ):
        nombre    = fila[0]
        sistema   = fila[1]
        conductor = fila[2]
        paralelos = fila[3]
        I_diseno  = fila[4]
        cos_phi   = fila[5]
        L_m       = fila[6]
        temp_amb  = fila[7]

        if nombre is None:
            continue

        sistema   = str(sistema).strip().upper()
        conductor = str(conductor).strip().upper()

        if sistema not in FACTOR_SISTEMA:
            errores.append(
                f"Fila {num_fila} '{nombre}': "
                f"sistema '{sistema}' inválido — usar 1F, 2F o 3F"
            )
            continue

        _tabla_validacion = {
            **get_tabla_conductores("AWG"),
            **get_tabla_conductores("MM2")
        }
        if conductor not in _tabla_validacion:
            errores.append(
                f"Fila {num_fila} '{nombre}': "
                f"conductor '{conductor}' no existe en tabla"
            )
            continue

        try:
            paralelos = int(paralelos)
            I_diseno  = float(I_diseno)
            cos_phi   = float(cos_phi)
            L_m       = float(L_m)
            temp_amb  = float(temp_amb)
        except (TypeError, ValueError):
            errores.append(
                f"Fila {num_fila} '{nombre}': "
                f"valor no numérico en columnas D-H"
            )
            continue

        if I_diseno <= 0:
            errores.append(f"Fila {num_fila} '{nombre}': corriente debe ser mayor a 0A")
            continue

        if not 0 < cos_phi <= 1:
            errores.append(f"Fila {num_fila} '{nombre}': factor de potencia debe estar entre 0 y 1")
            continue

        if L_m <= 0:
            errores.append(f"Fila {num_fila} '{nombre}': longitud debe ser mayor a 0 metros")
            continue

        if int(temp_amb) not in FACTORES_TEMP:
            errores.append(
                f"Fila {num_fila} '{nombre}': "
                f"temperatura {temp_amb}°C no está en tabla — usar 25, 30, 35, 40, 45 o 50"
            )
            continue

        circuitos.append({
            "nombre":    str(nombre).strip(),
            "sistema":   sistema,
            "conductor": conductor,
            "S_mm2":     CONDUCTORES[conductor]["mm2"],
            "I_max":     CONDUCTORES[conductor]["I_max"],
            "paralelos": paralelos,
            "I_diseno":  I_diseno,
            "cos_phi":   cos_phi,
            "L_m":       L_m,
            "temp_amb":  temp_amb,
        })

    if errores:
        print(f"\n  ADVERTENCIAS ({len(errores)} filas con problemas):")
        for e in errores:
            print(f"  ⚠ {e}")

    return circuitos


def enriquecer_circuitos(circuitos: list, norma: str = "AWG") -> list:
    """
    Resuelve S_mm2 e I_max usando la tabla correcta según norma.
    Debe llamarse después de leer_circuitos_excel().
    """
    tabla = get_tabla_conductores(norma)
    for c in circuitos:
        conductor = c["conductor"]
        if conductor in tabla:
            c["S_mm2"] = tabla[conductor]["mm2"]
            c["I_max"]  = tabla[conductor]["I_max"]
    return circuitos


# ============================================================
# LECTURA — PERFIL DE PROYECTO
# ============================================================

def leer_perfil_excel(libro_openpyxl):
    """
    Lee la hoja 'perfil' del libro Excel — modo híbrido.
    Estructura: columna A = campo, columna B = valor

    Campos:
        perfil             : domestico / comercial / industrial
        nombre_proyecto    : nombre del proyecto
        usar_transformador : si / no
        usar_protecciones  : si / no
        usar_balance       : si / no

    Retorna diccionario o None si no existe la hoja.
    """
    hoja_nombre = None
    for nombre in libro_openpyxl.sheetnames:
        if nombre.lower() == "perfil":
            hoja_nombre = nombre
            break

    if not hoja_nombre:
        return None

    hoja  = libro_openpyxl[hoja_nombre]
    datos = {}

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if not fila[0]:
            continue
        campo = str(fila[0]).strip().lower()
        # Normalizar campo — sin tildes
        for orig, repl in [("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u"),("ñ","n")]:
            campo = campo.replace(orig, repl)
        valor = str(fila[1]).strip().lower() if fila[1] is not None else ""
        datos[campo] = valor

    return {
        "perfil":             datos.get("perfil", "industrial"),
        "nombre_proyecto":    datos.get("nombre_proyecto", ""),
        "norma":              datos.get("norma", "AWG"),
        "usar_transformador": datos.get("usar_transformador", "si") == "si",
        "usar_protecciones":  datos.get("usar_protecciones", "si") == "si",
        "usar_balance":       datos.get("usar_balance", "si") == "si",
    }


# ============================================================
# EXPORTAR — TXT
# ============================================================

def guardar_txt(lineas, nombre_archivo):
    """Guarda lista de líneas en archivo .txt"""
    with open(nombre_archivo, "w", encoding="utf-8") as f:
        for linea in lineas:
            f.write(linea + "\n")
    print(f"  TXT guardado : {nombre_archivo}")

# ============================================================
# EXPORTAR — EXCEL FORMATEADO
# ============================================================

def exportar_excel(nombre_proyecto, circuitos, fecha, nombre_archivo, perfil=None):
    """
    Exporta resultados a Excel con formato profesional.
    Colores por estado: verde ÓPTIMO, amarillo ACEPTABLE,
    naranja PRECAUCIÓN, rojo FALLA.
    """
    perfil = perfil or {}
    COLORES = {
        "ÓPTIMO":     "FF92D050",
        "ACEPTABLE":  "FFFFFF00",
        "PRECAUCIÓN": "FFFFC000",
        "FALLA":      "FFFF0000",
        "SUPERA":     "FFFF0000",
    }

    borde = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    libro = openpyxl.Workbook()
    hoja  = libro.active
    hoja.title = "Resultados"

    # Título
    hoja.merge_cells("A1:L1")
    hoja["A1"] = f"REPORTE BT — {nombre_proyecto}"
    hoja["A1"].font      = Font(bold=True, size=13, color="FFFFFFFF")
    hoja["A1"].alignment = Alignment(horizontal="center")
    hoja["A1"].fill      = PatternFill("solid", fgColor="FF1F3864")

    # Info
    hoja.merge_cells("A2:L2")
    hoja["A2"] = f"Fecha: {fecha}  |  Normativa: SEC RIC N°10 / NEC / IEC 60364  |  Límite ΔV: {LIMITE_DV}%"
    hoja["A2"].alignment = Alignment(horizontal="center")
    hoja["A2"].fill      = PatternFill("solid", fgColor="FFD6E4F7")

    # Encabezados
    encabezados = [
        "Circuito", "Sistema", "Conductor", "Paralelos",
        "I_diseño(A)", "I_cap(A)", "Estado_I",
        "Potencia(W)", "ΔV(V)", "ΔV(%)", "Estado_dV", "Sugerencia"
    ]

    for col, texto in enumerate(encabezados, start=1):
        celda = hoja.cell(row=3, column=col, value=texto)
        celda.font      = Font(bold=True, color="FFFFFFFF")
        celda.fill      = PatternFill("solid", fgColor="FF2E75B6")
        celda.alignment = Alignment(horizontal="center")
        celda.border    = borde

    # Datos
    for i, c in enumerate(circuitos):
        fila_num = 4 + i

        I_cap        = capacidad_corregida(c["I_max"], c["paralelos"], c["temp_amb"])
        P_watts      = calcular_potencia(c["I_diseno"], c["cos_phi"], c["sistema"])
        dV_V, dV_pct = calcular_caida_tension(
            c["L_m"], c["S_mm2"], c["I_diseno"], c["paralelos"], c["sistema"]
        )
        estado_dV = clasificar_caida(dV_pct)
        estado_I  = "OK" if c["I_diseno"] <= I_cap else "SUPERA"

        sugerencia = ""
        if estado_dV == "FALLA" or estado_I == "SUPERA":
            cond, mm2, dv = sugerir_conductor(
                c["L_m"], c["I_diseno"], c["paralelos"],
                c["sistema"], c["temp_amb"],
                norma=perfil.get("norma", "AWG")
            )
            if cond:
                sugerencia = f"→ Usar {cond} ({dv}%)"

        desc_cond = f"{c['paralelos']}x{c['conductor']}" if c["paralelos"] > 1 else c["conductor"]

        valores = [
            c["nombre"], c["sistema"], desc_cond, c["paralelos"],
            c["I_diseno"], I_cap, estado_I,
            P_watts, dV_V, dV_pct, estado_dV, sugerencia
        ]

        for col, valor in enumerate(valores, start=1):
            celda = hoja.cell(row=fila_num, column=col, value=valor)
            celda.border    = borde
            celda.alignment = Alignment(horizontal="center")

        hoja.cell(row=fila_num, column=7).fill  = PatternFill("solid", fgColor=COLORES.get(estado_I,  "FFFFFFFF"))
        hoja.cell(row=fila_num, column=11).fill = PatternFill("solid", fgColor=COLORES.get(estado_dV, "FFFFFFFF"))

        if sugerencia:
            hoja.cell(row=fila_num, column=12).font = Font(bold=True, color="FFFF0000")

    # Anchos de columna
    anchos = [28, 8, 14, 10, 12, 10, 10, 12, 8, 8, 12, 22]
    for col, ancho in enumerate(anchos, start=1):
        hoja.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

    libro.save(nombre_archivo)
    print(f"  Excel guardado: {nombre_archivo}")
def leer_balance_excel(libro_openpyxl):
    """
    Lee la hoja 'balance' del libro Excel.
    Estructura: nombre | tablero | fase | tipo_carga
    Retorna dict indexado por nombre de circuito.
    """
    balance = {}

    hoja_nombre = None
    for nombre in libro_openpyxl.sheetnames:
        if nombre.lower() == "balance":
            hoja_nombre = nombre
            break

    if not hoja_nombre:
        return balance

    hoja = libro_openpyxl[hoja_nombre]

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if not fila[0]:
            continue
        nombre     = str(fila[0]).strip()
        tablero    = str(fila[1]).strip() if fila[1] else "SIN TABLERO"
        fase       = str(fila[2]).strip().upper() if fila[2] else "L1"
        tipo_carga = str(fila[3]).strip().lower() if fila[3] else "critica"

        balance[nombre] = {
            "tablero":    tablero,
            "fase":       fase,
            "tipo_carga": tipo_carga,
        }

    return balance


def leer_tableros_excel(libro_openpyxl):
    """
    Lee la hoja 'tableros' del libro Excel.
    Estructura: nombre | capacidad_kva
    Retorna dict {nombre_tablero: capacidad_kva}.
    """
    tableros = {}

    hoja_nombre = None
    for nombre in libro_openpyxl.sheetnames:
        if nombre.lower() == "tableros":
            hoja_nombre = nombre
            break

    if not hoja_nombre:
        return tableros

    hoja = libro_openpyxl[hoja_nombre]

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if not fila[0]:
            continue
        nombre   = str(fila[0]).strip()
        cap_kva  = float(fila[1]) if fila[1] else 0
        tableros[nombre] = cap_kva

    return tableros

def leer_demanda_excel(libro_openpyxl):
    """
    Lee la hoja 'demanda' del libro Excel — parámetros M6.
    Estructura: columna A = campo, columna B = valor

    Campos esperados:
        tipo_instalacion   : residencial/comercial/industrial/datacenter
        tipo_alimentador   : transformador/sec
        tension_alim       : float — V (220 o 380)
        sistema_alim       : 1F/3F
        cos_phi_global     : float — 0.0 a 1.0
        factor_crecimiento : float — ej: 1.25
        zona_sec           : urbana/suburbana/rural

    Retorna dict con parámetros normalizados, o None si no hay hoja.
    """
    hoja_nombre = None
    for nombre in libro_openpyxl.sheetnames:
        if nombre.lower() == "demanda":
            hoja_nombre = nombre
            break

    if not hoja_nombre:
        return None

    hoja  = libro_openpyxl[hoja_nombre]
    datos = {}

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if not fila[0]:
            continue
        campo = str(fila[0]).strip().lower()
        # Normalizar — sin tildes
        for orig, repl in [("á","a"),("é","e"),("í","i"),
                           ("ó","o"),("ú","u"),("ñ","n")]:
            campo = campo.replace(orig, repl)
        valor = str(fila[1]).strip() if fila[1] is not None else ""
        datos[campo] = valor

    # Construir resultado con tipos correctos y valores por defecto
    try:
        tension = float(datos.get("tension_alim", 380))
    except ValueError:
        tension = 380.0

    try:
        cos_phi = float(datos.get("cos_phi_global", 0.85))
        cos_phi = max(0.1, min(1.0, cos_phi))   # clamp 0.1-1.0
    except ValueError:
        cos_phi = 0.85

    try:
        f_crec = float(datos.get("factor_crecimiento", 1.0))
        f_crec = max(1.0, f_crec)               # mínimo 1.0
    except ValueError:
        f_crec = 1.0

    return {
        "tipo_instalacion":   datos.get("tipo_instalacion",  "industrial").lower(),
        "tipo_alimentador":   datos.get("tipo_alimentador",  "transformador").lower(),
        "tension_alim":       tension,
        "sistema_alim":       datos.get("sistema_alim",      "3F").upper(),
        "cos_phi_global":     cos_phi,
        "factor_crecimiento": f_crec,
        "zona_sec":           datos.get("zona_sec",          "urbana").lower(),
    }

def leer_cadena_excel(libro_openpyxl):
    """
    Lee la hoja 'cadena' del libro Excel — dispositivos para M7.

    Estructura: fila 1 puede ser título, encabezados en fila 1 o 2.
    Detecta automáticamente buscando la columna 'nombre'.

    Retorna lista de dicts, uno por dispositivo.
    Retorna lista vacía si no hay hoja 'cadena'.
    """
    hoja_nombre = None
    for nombre in libro_openpyxl.sheetnames:
        if nombre.lower() == "cadena":
            hoja_nombre = nombre
            break

    if not hoja_nombre:
        return []

    hoja = libro_openpyxl[hoja_nombre]

    # Detectar fila de encabezados — fila 1 o fila 2
    fila1 = [str(c.value).strip().lower() if c.value else "" for c in hoja[1]]
    fila2 = [str(c.value).strip().lower() if c.value else "" for c in hoja[2]]

    if "nombre" in fila2:
        encabezados   = fila2
        datos_min_row = 3
    elif "nombre" in fila1:
        encabezados   = fila1
        datos_min_row = 2
    else:
        return []

    dispositivos = []
    for fila in hoja.iter_rows(min_row=datos_min_row, values_only=True):
        if not any(v is not None for v in fila):
            continue

        d = {}
        for i, val in enumerate(fila):
            if i >= len(encabezados):
                break
            d[encabezados[i]] = val

        nombre_val = str(d.get("nombre") or "").strip()
        # Ignorar filas de notas/metadata (nombre muy largo o símbolo especial)
        if not nombre_val or len(nombre_val) > 40 or nombre_val.startswith("⚡"):
            continue

        def _float(key, default=None):
            v = d.get(key)
            if v is None or str(v).strip() in ("", "—", "-", "None"):
                return default
            try:
                return float(v)
            except (ValueError, TypeError):
                return default

        def _int(key, default=0):
            v = d.get(key)
            try:
                return int(float(v)) if v is not None else default
            except (ValueError, TypeError):
                return default

        def _str(key, default=""):
            v = d.get(key)
            return str(v).strip() if v is not None else default

        dispositivos.append({
            "nombre":       _str("nombre"),
            "designacion":  _str("designacion"),
            "circuito_ref": _str("circuito_ref"),
            "nivel":        _int("nivel", 0),
            "In_A":         _float("in_a", 0),
            "curva":        _str("curva", "C").upper(),
            "Ir_xIn":       _float("ir_xin",  1.0),
            "Isd_xIr":      _float("isd_xir", None),
            "tsd_s":        _float("tsd_s",   None),
            "Ii_xIn":       _float("ii_xin",  None),
            "modo":         _str("modo", "red").lower(),
            "upstream":     _str("upstream"),
            "Icc_kA":       _float("icc_ka",  None),
        })

    return dispositivos
