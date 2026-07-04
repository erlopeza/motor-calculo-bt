"""Robustez de la capa de entrada excel.py ante archivos de proyecto malformados.

Ningún lector debe crashear con datos basura/parciales: deben degradar con
gracia (None / dict vacío / omitir la fila problemática).
"""
import warnings
import openpyxl
import pytest

import excel

warnings.filterwarnings("ignore")


def _wb(tmp_path, sheets: dict) -> str:
    """sheets = {hoja: [filas]}. Escribe un .xlsx y retorna su ruta."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for nombre, filas in sheets.items():
        ws = wb.create_sheet(nombre)
        for fila in filas:
            ws.append(fila)
    if not wb.sheetnames:
        wb.create_sheet("vacia")
    ruta = str(tmp_path / "edge.xlsx")
    wb.save(ruta)
    return ruta


def _libro(tmp_path, sheets: dict):
    return openpyxl.load_workbook(_wb(tmp_path, sheets), data_only=True)


# ---------------------------------------------------------------------------
# Transformador
# ---------------------------------------------------------------------------

class TestTransformadorRobustez:
    # El lector salta la fila 1 (encabezado); los datos van desde la fila 2.
    _HDR = ["campo", "valor"]

    def test_kva_no_numerico_modo_A_no_crashea(self, tmp_path):
        p = _wb(tmp_path, {"Transformador": [
            self._HDR, ["modo", "A"], ["kva", "no-numerico"], ["vn_bt", 380], ["ucc_pct", 5]]})
        assert excel.leer_transformador_excel(p) is None

    def test_kva_no_numerico_modo_B_no_crashea(self, tmp_path):
        p = _wb(tmp_path, {"Transformador": [self._HDR, ["modo", "B"], ["kva", "abc"]]})
        assert excel.leer_transformador_excel(p) is None

    def test_vn_no_numerico_no_crashea(self, tmp_path):
        p = _wb(tmp_path, {"Transformador": [
            self._HDR, ["modo", "A"], ["kva", 1000], ["vn_bt", "###"], ["ucc_pct", 5]]})
        assert excel.leer_transformador_excel(p) is None

    def test_trafo_valido_sigue_funcionando(self, tmp_path):
        p = _wb(tmp_path, {"Transformador": [
            self._HDR, ["modo", "A"], ["kva", 1000], ["vn_bt", 380], ["ucc_pct", 5]]})
        r = excel.leer_transformador_excel(p)
        assert r is not None
        assert r["kVA"] == 1000.0 and r["modo"] == "A"


# ---------------------------------------------------------------------------
# Balance
# ---------------------------------------------------------------------------

class TestBalanceRobustez:
    def test_fila_corta_no_crashea(self, tmp_path):
        # fila con menos columnas de las esperadas (solo 2)
        lib = _libro(tmp_path, {"balance": [
            ["nombre", "tablero", "fase", "tipo_carga"], ["C-01", "TD-1"]]})
        r = excel.leer_balance_excel(lib)
        assert isinstance(r, dict)
        assert "C-01" in r  # la fila se leyó con defaults para lo faltante

    def test_valores_basura_no_crashea(self, tmp_path):
        lib = _libro(tmp_path, {"balance": [["campo", "valor"], ["kva", "xyz"]]})
        assert isinstance(excel.leer_balance_excel(lib), dict)

    def test_balance_valido(self, tmp_path):
        lib = _libro(tmp_path, {"balance": [
            ["nombre", "tablero", "fase", "tipo_carga"],
            ["C-01", "TD-1", "L1", "critica"]]})
        r = excel.leer_balance_excel(lib)
        assert r["C-01"]["tablero"] == "TD-1"


# ---------------------------------------------------------------------------
# Tableros
# ---------------------------------------------------------------------------

class TestTablerosRobustez:
    def test_capacidad_no_numerica_no_crashea(self, tmp_path):
        lib = _libro(tmp_path, {"tableros": [
            ["nombre", "capacidad_kva"], ["TD-1", "xyz"]]})
        r = excel.leer_tableros_excel(lib)
        assert isinstance(r, dict)
        assert r.get("TD-1") == 0.0  # no numérico → default 0

    def test_fila_corta_no_crashea(self, tmp_path):
        lib = _libro(tmp_path, {"tableros": [["nombre", "capacidad_kva"], ["TD-1"]]})
        r = excel.leer_tableros_excel(lib)
        assert isinstance(r, dict)

    def test_tableros_valido(self, tmp_path):
        lib = _libro(tmp_path, {"tableros": [
            ["nombre", "capacidad_kva"], ["TD-1", 250.0]]})
        assert excel.leer_tableros_excel(lib)["TD-1"] == 250.0
